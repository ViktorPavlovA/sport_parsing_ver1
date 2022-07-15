import requests
from bs4 import BeautifulSoup
import fake_useragent
import csv
import os
import shutil
import pandas as pd
import openpyxl

class data_parcer:
    # "https://russianblogs.com/article/1212876419/"
    def get_data(self, link_on_website):
        user_agent_new = fake_useragent.UserAgent().random
        dic_user_agent_new = {'user-agent': user_agent_new}
        responce = requests.get(link_on_website, dic_user_agent_new).text
        bs = BeautifulSoup(responce, 'html.parser')
        block = bs.find("div", id="branding-layout")
        find_score = block.find_all('b') #тег счета
        find_mistake = block.find_all('a',{"class": "score"})
        list_mistake =[]
        index_pre =[]
        try:
            for i in range(250):
                # print(f"mistake{find_mistake[i].text}")
                list_mistake.append(find_mistake[i].text)
                if list_mistake[i] == 'превью':
                    index_pre.append(i)

        except:pass
        # print(list_mistake)

        # print(f"index_preview{index_pre}") #индексы превью

        # print(f"mistake{find_mistake}")
        find_area = block.find_all("td", {"class": "alRight padR20"})
        area_list =[]
        try:
            for i in range(1, 250):  # исключение сюда
                area_list.append(block.find_all("td", {"class": "alRight padR20"})[i].text)
        except: pass

        # find_name = block.find_all('a')  # тег счета
        names_list =[]
        try:
            for i in range(1, 250):  # исключение сюда
                if i % 2 == 0:
                    pass
                else:
                    names_list.append(block.find_all("div", {"class": "hide-field"})[i].text)
        except: pass
        x =[]
        names_list = [x.replace('\xa0', '') for x in names_list] #чтобы удалить из всех элементов \xa0
        list_final_score = []
        for i in find_score:
            list_final_score.append(i.text)
        #счетчик индексов элементов от, б
        defect_ind =[]
        clean_score =[]
        try:
            for i in range(350):
                if list_final_score[i] in ("от ",' от','б ',' б'):
                    defect_ind.append(i)
                else: clean_score.append(list_final_score[i])
        except:pass
        list_final_score.clear()
        list_final_score = clean_score
        del clean_score

        clean_name_list = []
        for i in range(len(names_list)):
            if i in index_pre:pass
            else: clean_name_list.append(names_list[i])
        names_list.clear()
        names_list =clean_name_list
        del clean_name_list

        clean_area_list = []
        for i in range(len(area_list)):
            if i in index_pre:
                pass
            else:
                clean_area_list.append(area_list[i])
        area_list.clear()
        area_list = clean_area_list
        del clean_area_list

        # print(f"list_final_score:{len(list_final_score)}{list_final_score}\n"
        #       f"list_final_score:{len(names_list)}{names_list}\n"
        #       f"list_final_score:{len(area_list)}{area_list}\n")

        return list_final_score,names_list,area_list,index_pre

    def divide(self, list_score,index_pre):
        list_divided =[]
        clean_list = []
        # print(list_score)
        for i in list(list_score):
            for y in list(i):
                list_divided.append(y)
        # print(f"list_score_in_def{list_divided}")
        for i in range(len(list_divided)):
            if list_divided[i] in (' ',':','о','т','б'): #удаление символов из списка
                pass
            else:
                clean_list.append(list_divided[i])
        return clean_list


    def save_xlsx(self,list_score,name_list,area_list,team_list,location_file,j):
        def function2writelists( list_score, name_list, area_list):
            wb = openpyxl.load_workbook(r'data/sport_data_test.xlsx')
            #main_body
            ws = wb.active
            for i in range(len(name_list)):
                ws['A{}'.format(i+2)] = name_list[i]
                ws['B{}'.format(i + 2)] = area_list[i]
                if area_list[i]== 'В гостях':
                    ws['C{}'.format(i + 2)] = int(list_score[2*i+1])
                    ws['D{}'.format(i + 2)] = int(list_score[2*i])
                elif area_list[i]== 'Дома':
                    ws['C{}'.format(i + 2)] = int(list_score[2*i])
                    ws['D{}'.format(i + 2)] = int(list_score[2*i+1])


            wb.save(r'data/sport_data_test.xlsx')
        if (j>0): #creat and jump to new page
            wb = openpyxl.load_workbook(r'data/sport_data_test.xlsx')
            wb.create_sheet(title=team_list)
            wb.active = j
            ws = wb.active
            ws['A1'] = "Соперник"
            ws['B1'] = "Арена"
            ws['C1'] = "Счет {}".format(team_list)
            ws['D1'] = "Счет команды соперника"
            wb.save(r'data/sport_data_test.xlsx')
            function2writelists(list_score, name_list, area_list)

        elif j == 0: #for first time
            flag_mk = os.path.exists("data")
            os.mkdir("data")
            wb = openpyxl.Workbook()
            sheets = wb.sheetnames
            sheet_name = wb[sheets[j]]
            sheet_name.title = '{}'.format(team_list)
            ws = wb.active
            ws['A1'] = "Соперник"
            ws['B1'] = "Арена"
            ws['C1'] = "Счет {}".format(team_list)
            ws['D1'] = "Счет команды соперника"
            wb.save(r'data/sport_data_test.xlsx')
            function2writelists(list_score,name_list,area_list)

    def del_folder(self):
        shutil.rmtree("data")

    def make_a_big_list(self,list_old,list_new):
        list_old=list_old+list_new
        return list_old


if __name__ == '__main__':
    location_file ="data/sport_data_test.csv"
    teams_list = ["washington-capitals","new-york-rangers","tampa-bay-lightning","colorado-avalanche",
                  "new-jersey-devils","calgary-flames","ottawa-senators","detroit-red-wings","phoenix-coyotes",
                  "philadelphia-flyers","florida-panthers","columbus-blue-jackets","anaheim-ducks","los-angeles-kings",
                  "boston-bruins","toronto-maple-leafs","edmonton-oilers","st-louis-blues","minnesota-wild",
                  "vancouver-canucks","dallas-stars","buffalo-sabres","winnipeg-jets","chicago-blackhawks"]
    years_list =[2014,2015,2016,2017,2018,2019,2020,2021,2022]
    link_on_your_tor = 'C:/Users/Виктор/Documents/Tor Browser/Browser/firefox.exe'
    open_port = 9150
    x = data_parcer
    big_list_score = []
    big_list_name = []
    big_list_area = []
    try:
        x.del_folder(None)  # < -- if you want delete folder data
    except:
        pass
    for j in range(len(teams_list)):
        big_list_score.clear()
        big_list_name.clear()
        big_list_area.clear()
        print(f'Right now team is {teams_list[j]}')
        for i in range(len(years_list) - 1):
            link_on_website = "https://www.sports.ru/{}/calendar/{}-{}/".format(teams_list[j], str(years_list[i]),
                                                                                str(years_list[i + 1]))
            print(f"{str(years_list[i])}\t{str(years_list[i + 1])}")
            list_score, name_list, area_list, index_pre = x.get_data(None, link_on_website)
            list_score = x.divide(None, list_score, index_pre)
            big_list_score =x.make_a_big_list(None,big_list_score,list_score)
            big_list_name = x.make_a_big_list(None,big_list_name,name_list)
            big_list_area = x.make_a_big_list(None,big_list_area,area_list)
            if i == len(years_list)-2: #на этом шаге нужно передать все значения для сохранения в xlsx
                x.save_xlsx(None, big_list_score, big_list_name, big_list_area, teams_list[j], location_file,
                            j)  # < -- if you want save xlsx
                print(big_list_score,len(big_list_score))
                print(big_list_name,len(big_list_name))
                print(big_list_area, len(big_list_area))
